"""
Excel report generator using OpenPyXL.
Produces a multi-sheet, audit-grade reconciliation workbook.
"""
from __future__ import annotations

import io
from datetime import datetime, date
from typing import Any

import structlog
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

logger = structlog.get_logger(__name__)

# ── Color palette ─────────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1E3A5F"   # Dark navy
CLR_HEADER_FG   = "FFFFFF"
CLR_MATCH       = "D4EDDA"   # Light green
CLR_MISMATCH    = "F8D7DA"   # Light red
CLR_WARNING     = "FFF3CD"   # Amber
CLR_DUPLICATE   = "FCE4EC"   # Pink
CLR_CRITICAL    = "B71C1C"   # Deep red (text)
CLR_ALT_ROW     = "F5F7FA"   # Light grey
CLR_BORDER      = "DEE2E6"
CLR_BRAND       = "3B82F6"   # Recko blue

INR = '#,##0.00'


def _header_font() -> Font:
    return Font(bold=True, color=CLR_HEADER_FG, name="Calibri", size=11)

def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=CLR_HEADER_BG)

def _alt_fill() -> PatternFill:
    return PatternFill("solid", fgColor=CLR_ALT_ROW)

def _mismatch_fill() -> PatternFill:
    return PatternFill("solid", fgColor=CLR_MISMATCH)

def _match_fill() -> PatternFill:
    return PatternFill("solid", fgColor=CLR_MATCH)

def _thin_border() -> Border:
    side = Side(style="thin", color=CLR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _money(ws: Any, cell: Any, value: float) -> None:
    cell.value = round(float(value or 0), 2)
    cell.number_format = INR

def _style_header_row(ws: Any, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _center()
        cell.border = _thin_border()

def _auto_width(ws: Any, min_width: int = 12, max_width: int = 40) -> None:
    for col in ws.columns:
        length = max(
            len(str(cell.value or "")) for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            min_width, min(length + 4, max_width)
        )

def _freeze_header(ws: Any) -> None:
    ws.freeze_panes = "A2"


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_summary_sheet(wb: Workbook, run_data: dict[str, Any]) -> None:
    ws = wb.create_sheet("1. Executive Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws["B2"] = "Recko GST Reconciliation Report"
    ws["B2"].font = Font(bold=True, size=18, color=CLR_BRAND, name="Calibri")
    ws["B3"] = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws["B3"].font = Font(size=10, color="888888", name="Calibri")

    # Summary table
    data = [
        ("Period", run_data.get("return_period", "N/A")),
        ("Financial Year", run_data.get("fy", "N/A")),
        ("Total PR Records", run_data.get("total_pr_records", 0)),
        ("Total GSTR-2B Records", run_data.get("total_2b_records", 0)),
        ("Matched Records", run_data.get("matched_count", 0)),
        ("Unmatched (PR)", run_data.get("unmatched_pr_count", 0)),
        ("Unmatched (2B)", run_data.get("unmatched_2b_count", 0)),
        ("Duplicates Detected", run_data.get("duplicate_count", 0)),
        ("Total ITC Claimed (₹)", run_data.get("total_itc_claimed", 0)),
        ("ITC at Risk (₹)", run_data.get("itc_at_risk", 0)),
        ("ITC Matched (₹)", run_data.get("itc_matched", 0)),
    ]

    start_row = 6
    for i, (label, value) in enumerate(data):
        row = start_row + i
        label_cell = ws.cell(row=row, column=2, value=label)
        val_cell = ws.cell(row=row, column=3, value=value)

        label_cell.font = Font(bold=True, name="Calibri", size=11)
        label_cell.border = _thin_border()
        val_cell.border = _thin_border()

        if "₹" in label:
            _money(ws, val_cell, float(value))  # type: ignore[arg-type]

        if i % 2 == 0:
            for c in [label_cell, val_cell]:
                c.fill = _alt_fill()

        if "at Risk" in label:
            val_cell.font = Font(bold=True, color=CLR_CRITICAL, name="Calibri")

    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22


def _build_mismatches_sheet(
    wb: Workbook, mismatches: list[dict[str, Any]]
) -> None:
    ws = wb.create_sheet("2. Mismatches")
    ws.freeze_panes = "A2"

    headers = [
        "Source", "Invoice No", "GSTIN Supplier", "Supplier Name",
        "Invoice Date", "Taxable Value (₹)", "IGST (₹)", "CGST (₹)", "SGST (₹)",
        "ITC Impact (₹)", "Category", "Mismatch Fields", "Status",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    CATEGORY_COLORS = {
        "MISSING_IN_2B":           CLR_MISMATCH,
        "GSTIN_MISMATCH":          CLR_MISMATCH,
        "AMOUNT_VARIANCE":         CLR_WARNING,
        "TAX_RATE_MISMATCH":       CLR_WARNING,
        "DUPLICATE_INVOICE":       CLR_DUPLICATE,
        "MISSING_IN_PR":           CLR_WARNING,
        "DATE_MISMATCH":           CLR_ALT_ROW,
        "INVOICE_NUMBER_MISMATCH": CLR_ALT_ROW,
    }

    for i, rec in enumerate(mismatches, start=2):
        category = str(rec.get("category", ""))
        fill_color = CATEGORY_COLORS.get(category, CLR_ALT_ROW)
        row_fill = PatternFill("solid", fgColor=fill_color)

        row = [
            rec.get("source"),
            rec.get("invoice_number"),
            rec.get("gstin_supplier"),
            rec.get("supplier_name"),
            rec.get("invoice_date"),
            round(float(rec.get("taxable_value", 0) or 0), 2),
            round(float(rec.get("igst", 0) or 0), 2),
            round(float(rec.get("cgst", 0) or 0), 2),
            round(float(rec.get("sgst", 0) or 0), 2),
            round(float(rec.get("itc_impact", 0) or 0), 2),
            category,
            ", ".join(rec.get("mismatch_fields", []) or []),
            rec.get("status", "open"),
        ]
        ws.append(row)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            cell.fill = row_fill
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center")

        # Format money columns
        for col_idx in [6, 7, 8, 9, 10]:
            ws.cell(row=i, column=col_idx).number_format = INR

    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


def _build_duplicates_sheet(
    wb: Workbook, duplicates: list[dict[str, Any]]
) -> None:
    ws = wb.create_sheet("3. Duplicates")
    ws.freeze_panes = "A2"

    headers = [
        "Source", "Invoice Number", "GSTIN Supplier",
        "Duplicate Type", "Similarity Score", "Status",
        "Record Hash A", "Record Hash B",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for i, rec in enumerate(duplicates, start=2):
        row_fill = _alt_fill() if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row = [
            rec.get("source"),
            rec.get("invoice_number"),
            rec.get("gstin_supplier"),
            rec.get("dtype"),
            round(float(rec.get("similarity_score", 0) or 0), 4),
            rec.get("status"),
            rec.get("record_id_a", "")[:16] + "...",
            rec.get("record_id_b", "")[:16] + "...",
        ]
        ws.append(row)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            cell.fill = row_fill
            cell.border = _thin_border()

    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


def _build_vendor_sheet(
    wb: Workbook, vendor_stats: list[dict[str, Any]]
) -> None:
    ws = wb.create_sheet("4. Vendor Risk Analysis")
    ws.freeze_panes = "A2"

    headers = [
        "GSTIN", "Vendor Name", "PR Invoices", "GSTR-2B Invoices",
        "Matched", "Mismatched", "ITC Claimed (₹)", "ITC Matched (₹)",
        "ITC at Risk (₹)", "Mismatch Rate", "Risk Level",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    RISK_COLORS = {
        "critical": "FFCDD2",
        "high":     "FFE0B2",
        "medium":   "FFF9C4",
        "low":      "C8E6C9",
    }

    for i, vs in enumerate(vendor_stats, start=2):
        risk = str(vs.get("risk_level", "low")).lower()
        fill_color = RISK_COLORS.get(risk, CLR_ALT_ROW)
        row_fill = PatternFill("solid", fgColor=fill_color)

        row = [
            vs.get("vendor_gstin") or vs.get("gstin"),
            vs.get("vendor_name") or vs.get("name"),
            vs.get("pr_invoices", 0),
            vs.get("gstr2b_invoices", 0),
            vs.get("matched_invoices", 0),
            vs.get("mismatched_invoices", 0),
            round(float(vs.get("itc_claimed", 0) or 0), 2),
            round(float(vs.get("itc_matched", 0) or 0), 2),
            round(float(vs.get("itc_at_risk", 0) or 0), 2),
            f"{round(float(vs.get('mismatch_rate', 0) or 0) * 100, 1)}%",
            risk.upper(),
        ]
        ws.append(row)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            cell.fill = row_fill
            cell.border = _thin_border()

        for col_idx in [7, 8, 9]:
            ws.cell(row=i, column=col_idx).number_format = INR

    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


# ── Public API ────────────────────────────────────────────────────────────────

def generate_excel_report(
    run_data: dict[str, Any],
    mismatches: list[dict[str, Any]],
    duplicates: list[dict[str, Any]],
    vendor_stats: list[dict[str, Any]],
) -> bytes:
    """
    Generate a multi-sheet Excel reconciliation report.

    Args:
        run_data:      Reconciliation run summary dict
        mismatches:    List of unmatched_records dicts
        duplicates:    List of duplicate_records dicts
        vendor_stats:  List of vendor_run_stats dicts

    Returns:
        Raw bytes of the .xlsx file
    """
    logger.info(
        "report.excel.generating",
        run_id=run_data.get("id"),
        mismatches=len(mismatches),
        duplicates=len(duplicates),
        vendors=len(vendor_stats),
    )

    wb = Workbook()
    # Remove default blank sheet
    wb.remove(wb.active)  # type: ignore[arg-type]

    _build_summary_sheet(wb, run_data)
    _build_mismatches_sheet(wb, mismatches)
    _build_duplicates_sheet(wb, duplicates)
    _build_vendor_sheet(wb, vendor_stats)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    content = buf.read()

    logger.info("report.excel.done", size_bytes=len(content))
    return content
