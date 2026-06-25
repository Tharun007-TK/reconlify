import io
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class ExcelReportGenerator:
    """Generates multi-sheet Excel reports from reconciliation data."""

    def __init__(self, data: dict[str, Any]):
        self.data = data
        self.wb = Workbook()

    def generate(self) -> bytes:
        self._build_summary_sheet()
        self._build_matched_sheet()
        self._build_unmatched_sheet()
        self._build_duplicates_sheet()
        self._build_vendor_sheet()
        
        # Remove default empty sheet
        if "Sheet" in self.wb.sheetnames and len(self.wb.sheetnames) > 1:
            del self.wb["Sheet"]

        # Save to bytes buffer
        buf = io.BytesIO()
        self.wb.save(buf)
        return buf.getvalue()

    def _style_header(self, ws, row_num: int = 1) -> None:
        """Apply bold styling and a background color to a header row."""
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_adjust_columns(self, ws) -> None:
        """Adjust column widths based on the maximum length of the contents."""
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = min(adjusted_width, 50)

    def _build_summary_sheet(self) -> None:
        ws = self.wb.active
        ws.title = "Summary"
        run = self.data.get("run", {})

        ws.append(["Reconciliation Run Summary"])
        ws["A1"].font = Font(size=14, bold=True)
        
        ws.append([])
        
        metrics = [
            ("Status", run.get("status", "Unknown").upper()),
            ("Engine", run.get("engine_name", "N/A")),
            ("Started At", str(run.get("started_at", ""))),
            ("Completed At", str(run.get("completed_at", ""))),
            ("Matched Records", run.get("matched_count", 0)),
            ("Unmatched (PR)", run.get("unmatched_pr_count", 0)),
            ("Unmatched (2B)", run.get("unmatched_2b_count", 0)),
            ("Duplicates", run.get("duplicate_count", 0)),
            ("Match Rate", f"{run.get('match_rate', 0.0) * 100:.2f}%"),
            ("Total ITC Claimed", f"₹{run.get('total_itc_claimed', 0.0):,.2f}"),
            ("ITC Matched", f"₹{run.get('itc_matched', 0.0):,.2f}"),
            ("ITC At Risk", f"₹{run.get('itc_at_risk', 0.0):,.2f}"),
        ]

        for key, value in metrics:
            ws.append([key, value])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

        self._auto_adjust_columns(ws)

    def _build_matched_sheet(self) -> None:
        ws = self.wb.create_sheet(title="Matched Records")
        headers = [
            "Supplier GSTIN", "PR Invoice No", "PR Date", "PR Taxable Value", "PR Tax",
            "GSTR-2B Invoice No", "GSTR-2B Date", "GSTR-2B Taxable Value", "GSTR-2B Tax"
        ]
        ws.append(headers)
        self._style_header(ws)

        for record in self.data.get("matched", []):
            pr_tax = sum(filter(None, [record.get("pr_igst"), record.get("pr_cgst"), record.get("pr_sgst")]))
            b2_tax = sum(filter(None, [record.get("b2_igst"), record.get("b2_cgst"), record.get("b2_sgst")]))

            ws.append([
                record.get("gstin_supplier"),
                record.get("pr_invoice_number"),
                str(record.get("pr_invoice_date", "")),
                record.get("pr_taxable_value"),
                pr_tax,
                record.get("b2_invoice_number"),
                str(record.get("b2_invoice_date", "")),
                record.get("b2_taxable_value"),
                b2_tax,
            ])
        self._auto_adjust_columns(ws)

    def _build_unmatched_sheet(self) -> None:
        ws = self.wb.create_sheet(title="Unmatched Records")
        headers = [
            "Source", "Severity", "Category", "Supplier GSTIN", "Invoice Number",
            "Date", "Taxable Value", "Total Tax", "Reason", "Recommended Action"
        ]
        ws.append(headers)
        self._style_header(ws)

        for r in self.data.get("unmatched", []):
            tax = sum(filter(None, [r.get("igst"), r.get("cgst"), r.get("sgst")]))
            ws.append([
                r.get("source"),
                r.get("severity"),
                r.get("category"),
                r.get("gstin_supplier"),
                r.get("invoice_number"),
                str(r.get("invoice_date", "")),
                r.get("taxable_value"),
                tax,
                r.get("reason"),
                r.get("recommended_action")
            ])
            
            # Color code severity
            sev = r.get("severity", "")
            if sev == "Critical":
                ws.cell(row=ws.max_row, column=2).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif sev == "High":
                ws.cell(row=ws.max_row, column=2).fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
            elif sev == "Medium":
                ws.cell(row=ws.max_row, column=2).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

        self._auto_adjust_columns(ws)

    def _build_duplicates_sheet(self) -> None:
        ws = self.wb.create_sheet(title="Duplicates")
        headers = [
            "Source", "Type", "Status", "Similarity Score",
            "Supplier GSTIN", "Invoice Number", "Reason"
        ]
        ws.append(headers)
        self._style_header(ws)

        for d in self.data.get("duplicates", []):
            diffs = d.get("diff_fields", {})
            reason = diffs.get("reason", str(diffs)) if isinstance(diffs, dict) else str(diffs)
            
            ws.append([
                d.get("source"),
                d.get("dtype"),
                d.get("status"),
                f"{d.get('similarity_score', 0) * 100:.2f}%",
                d.get("gstin_supplier"),
                d.get("invoice_number"),
                reason
            ])
        self._auto_adjust_columns(ws)

    def _build_vendor_sheet(self) -> None:
        ws = self.wb.create_sheet(title="Vendor Analysis")
        headers = [
            "Risk Level", "GSTIN", "Name", "Mismatch Rate",
            "PR Invoices", "GSTR-2B Invoices", "Matched", "Mismatched",
            "ITC Claimed", "ITC Matched", "ITC At Risk"
        ]
        ws.append(headers)
        self._style_header(ws)

        for v in self.data.get("vendors", []):
            ws.append([
                v.get("risk_level", "").upper(),
                v.get("gstin"),
                v.get("name"),
                f"{v.get('mismatch_rate', 0) * 100:.2f}%",
                v.get("pr_invoices"),
                v.get("gstr2b_invoices"),
                v.get("matched_invoices"),
                v.get("mismatched_invoices"),
                v.get("itc_claimed"),
                v.get("itc_matched"),
                v.get("itc_at_risk")
            ])
            
            risk = v.get("risk_level", "")
            if risk == "critical":
                ws.cell(row=ws.max_row, column=1).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif risk == "high":
                ws.cell(row=ws.max_row, column=1).fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")

        self._auto_adjust_columns(ws)
