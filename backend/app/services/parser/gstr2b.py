"""
GSTR-2B Parser — Production Implementation.

Supports two official GST Portal download formats:
  1. JSON export  (data/docdata/b2b/inv[]/itms[]/itm_det structure)
  2. Excel export (portal-generated .xlsx / .xls / .csv spreadsheet)

A file-type detector routes to the correct parser based on MIME type or
file extension.  Both paths produce the same flat Pandas DataFrame with
canonical column names matching the NormalizedRecord schema.

JSON structure reference (GSTN GSTR-2B API v1.1):
  {
    "data": {
      "rtnprd": "032024",
      "gstin": "...",
      "docdata": {
        "b2b": [
          {
            "ctin":  "<Supplier GSTIN>",
            "trdnm": "<Trade Name>",
            "inv": [
              {
                "inum": "<Invoice Number>",
                "dt":   "<DD-MM-YYYY>",
                "val":  <Total invoice value>,
                "pos":  "<Place of Supply>",
                "rev":  "<Reverse Charge Y/N>",
                "itms": [
                  {
                    "num": <line item no>,
                    "itm_det": {
                      "rt":    <GST Rate %>,
                      "txval": <Taxable Value>,
                      "igst":  <IGST Amount>,
                      "cgst":  <CGST Amount>,
                      "sgst":  <SGST Amount>,
                      "cess":  <Cess Amount>
                    }
                  }
                ]
              }
            ]
          }
        ],
        "cdn": [...],   # Credit / Debit Notes
        "impg": [...]   # IGST on Imports
      }
    }
  }
"""
from __future__ import annotations

import hashlib
import io
import json
import mimetypes
from datetime import date, datetime
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
import structlog

from app.core.exceptions import ParseError
from app.services.parser.purchase_register import (
    ParseResult,
    _coerce_numeric,
    _normalize_columns,
    _validate_gstin,
)
from app.services.parser.schemas import (
    ColumnMapping,
    DocumentType,
    FileType,
    NormalizedRecord,
    ParseIssue,
    ParseSeverity,
    ParseResult as SchemaParseResult,
)

logger = structlog.get_logger(__name__)


# ── Column aliases for Excel / CSV portal downloads ────────────────────────────

GSTR2B_EXCEL_ALIASES: dict[str, list[str]] = {
    "gstin_supplier": [
        "gstin of supplier", "supplier gstin", "ctin", "gstin",
        "gstin_supplier", "vendor gstin", "party gstin",
    ],
    "supplier_name": [
        "trade/legal name of the supplier", "trade name", "supplier name",
        "trdnm", "vendor name", "party name", "supplier_name",
    ],
    "invoice_number": [
        "invoice number", "invoice no", "document number", "doc no",
        "inum", "invoice_number", "bill no",
    ],
    "invoice_date": [
        "invoice date", "document date", "date of invoice", "dt",
        "invoice_date", "bill date",
    ],
    "taxable_value": [
        "taxable value", "taxable amount", "value of supply",
        "txval", "taxable_value", "net amount",
    ],
    "igst": ["igst", "integrated tax", "igst amount", "igst_amount"],
    "cgst": ["cgst", "central tax", "cgst amount", "cgst_amount"],
    "sgst": ["sgst", "state tax", "sgst amount", "sgst_amount", "utgst"],
    "cess": ["cess", "cess amount", "cess_amount"],
    "document_type": [
        "document type", "invoice type", "supply type",
        "typ", "document_type",
    ],
    "itc_availability": [
        "itc availability", "itc eligible", "eligibility for itc",
        "itc_availability",
    ],
    "return_period": [
        "return period", "filing period", "return_period", "rtnprd",
    ],
    "place_of_supply": [
        "place of supply", "pos", "place_of_supply",
    ],
    "reverse_charge": [
        "reverse charge", "rev", "reverse_charge",
    ],
}

# MIME types that indicate JSON content
_JSON_MIMES = {
    "application/json",
    "text/json",
    "text/plain",   # Portal sometimes returns JSON with text/plain
}

# MIME types that indicate Excel/CSV content
_EXCEL_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # xlsx
    "application/vnd.ms-excel",                                            # xls
    "text/csv",
    "application/csv",
}

# Files above this threshold are streamed via openpyxl read_only=True
# to avoid loading the full workbook into memory at once.
_LARGE_FILE_THRESHOLD_BYTES: int = 5 * 1024 * 1024  # 5 MB


# ── File type detection ────────────────────────────────────────────────────────

def _detect_format(filename: str, mime_type: str | None = None) -> str:
    """
    Determine file format from MIME type (preferred) or extension (fallback).

    Returns one of: "json", "xlsx", "xls", "csv"

    Raises:
        ParseError: If the format is not supported.
    """
    # 1. Check MIME type first (most reliable)
    if mime_type:
        mt = mime_type.lower().split(";")[0].strip()
        if mt in _JSON_MIMES:
            # Double-check: a .xlsx with text/plain would be caught below by extension
            ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
            if ext in ("xlsx", "xls", "csv"):
                # Trust extension over generic MIME
                return ext
            return "json"
        if mt in _EXCEL_MIMES:
            if "csv" in mt:
                return "csv"
            ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
            return ext if ext in ("xlsx", "xls") else "xlsx"

    # 2. Fall back to file extension
    if "." in filename:
        ext = filename.lower().rsplit(".", 1)[-1]
        if ext == "json":
            return "json"
        if ext in ("xlsx", "xls", "csv"):
            return ext

    raise ParseError(
        f"Cannot determine GSTR-2B file format from filename='{filename}' "
        f"mime_type='{mime_type}'. Supported formats: JSON, XLSX, XLS, CSV."
    )


# ── Date parsing ───────────────────────────────────────────────────────────────

def _parse_gst_date(val: Any) -> date | None:
    """
    Parse a date value from any format encountered in GSTR-2B files.

    The JSON portal export uses DD-MM-YYYY.
    Excel exports use various locale-dependent formats.
    """
    if val is None:
        return None

    # Already a date / Timestamp
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, pd.Timestamp):
        return val.date()

    raw = str(val).strip()
    if not raw or raw.lower() in ("nan", "none", "null", "n/a", ""):
        return None

    # Try DD-MM-YYYY first (canonical GSTN JSON format)
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y", "%d/%m/%y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue

    # Pandas multi-format fallback (handles ISO, US, etc.)
    try:
        return pd.to_datetime(raw, dayfirst=True).date()
    except Exception:
        return None


# ── Helpers ────────────────────────────────────────────────────────────────────

def _compute_2b_row_hash(row: dict[str, Any], salt: str) -> str:
    key = "|".join([
        str(row.get("gstin_supplier", "")).upper().strip(),
        str(row.get("invoice_number", "")).upper().strip(),
        str(row.get("invoice_date", "")),
        str(round(row.get("taxable_value", 0.0), 2)),
        str(row.get("document_type", "INV")).upper(),
        salt,
    ])
    return hashlib.sha256(key.encode("utf-8")).hexdigest()


def _sum_items(items: list[dict[str, Any]], key: str) -> float:
    """Sum a numeric key across all items in an invoice line-item list."""
    return round(sum(_coerce_numeric(i.get(key, 0)) for i in items), 2)


# ── JSON parser ────────────────────────────────────────────────────────────────

def _parse_gstr2b_json(
    raw_bytes: bytes,
    run_id: str,
    client_id: str,
    salt: str,
) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    """
    Parse an official GSTR-2B JSON portal export.

    Navigates:
      data["docdata"]["b2b"]          → B2B invoices
      data["docdata"]["cdn"]          → Credit/Debit notes (optional)
      data["docdata"]["impg"]         → IGST on imports (metadata only)

    For each supplier in b2b[]:
      - ctin  → gstin_supplier
      - trdnm → supplier_name
      For each invoice in inv[]:
        - inum → invoice_number
        - dt   → invoice_date  (DD-MM-YYYY → datetime.date)
        - val  → invoice total (informational; taxable_value comes from itm_det)
        Aggregate across itms[].itm_det:
          txval → taxable_value
          igst  → igst
          cgst  → cgst
          sgst  → sgst
          cess  → cess

    Returns:
        (DataFrame with canonical columns, list of error dicts)
    """
    try:
        data = json.loads(raw_bytes)
    except json.JSONDecodeError as exc:
        raise ParseError(f"Invalid JSON in GSTR-2B file: {exc}") from exc

    # Navigate to the document data root
    root = data.get("data", data)
    return_period = str(root.get("rtnprd", "")).strip() or None
    doc_data = root.get("docdata", {})

    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    row_num = 1  # 1-indexed, no header row in JSON

    # ── B2B invoices ─────────────────────────────────────────────────────────
    for supplier in doc_data.get("b2b", []):
        gstin = str(supplier.get("ctin", "")).strip().upper()
        supplier_name = str(supplier.get("trdnm", "")).strip() or None

        if not gstin:
            errors.append({
                "row": row_num,
                "reason": "Supplier entry missing 'ctin' (GSTIN) — skipped",
                "raw": supplier,
            })
            continue

        for inv in supplier.get("inv", []):
            invoice_number = str(inv.get("inum", "")).strip().upper()
            invoice_date   = _parse_gst_date(inv.get("dt"))

            if not invoice_number:
                errors.append({
                    "row": row_num,
                    "gstin": gstin,
                    "reason": "Invoice missing 'inum' (invoice number) — skipped",
                })
                row_num += 1
                continue

            # Flatten itms[].itm_det → aggregate amounts
            items: list[dict[str, Any]] = [
                item.get("itm_det", item)
                for item in inv.get("itms", [])
            ]
            if not items:
                # Fallback: use inv-level totals if itms is absent
                items = [{"txval": inv.get("val", 0), "igst": 0, "cgst": 0, "sgst": 0, "cess": 0}]

            taxable_value = _sum_items(items, "txval")
            igst          = _sum_items(items, "igst")
            cgst          = _sum_items(items, "cgst")
            sgst          = _sum_items(items, "sgst")
            cess          = _sum_items(items, "cess")

            record: dict[str, Any] = {
                "run_id":         run_id,
                "client_id":      client_id,
                "source_row":     row_num,
                "gstin_supplier": gstin,
                "supplier_name":  supplier_name,
                "invoice_number": invoice_number,
                "invoice_date":   invoice_date,
                "taxable_value":  taxable_value,
                "igst":           igst,
                "cgst":           cgst,
                "sgst":           sgst,
                "cess":           cess,
                "document_type":  DocumentType.INVOICE,
                "is_amended":     False,
                "itc_availability": None,
                "return_period":  return_period,
                "place_of_supply": str(inv.get("pos", "")).strip() or None,
                "reverse_charge": str(inv.get("rev", "N")).strip().upper() == "Y",
            }
            record["row_hash"] = _compute_2b_row_hash(record, salt)
            rows.append(record)
            row_num += 1

    # ── Credit / Debit Notes ──────────────────────────────────────────────────
    for supplier in doc_data.get("cdn", []):
        gstin = str(supplier.get("ctin", "")).strip().upper()
        supplier_name = str(supplier.get("trdnm", "")).strip() or None

        for note in supplier.get("nt", []):
            note_num  = str(note.get("ntnum", "")).strip().upper()
            note_date = _parse_gst_date(note.get("ntdt"))
            note_type = str(note.get("typ", "C")).strip().upper()
            doc_type  = DocumentType.CREDIT_NOTE if note_type == "C" else DocumentType.DEBIT_NOTE

            items = [
                item.get("itm_det", item)
                for item in note.get("itms", [])
            ]
            if not items:
                items = [{"txval": 0, "igst": 0, "cgst": 0, "sgst": 0, "cess": 0}]

            record = {
                "run_id":          run_id,
                "client_id":       client_id,
                "source_row":      row_num,
                "gstin_supplier":  gstin,
                "supplier_name":   supplier_name,
                "invoice_number":  note_num,
                "invoice_date":    note_date,
                "taxable_value":   _sum_items(items, "txval"),
                "igst":            _sum_items(items, "igst"),
                "cgst":            _sum_items(items, "cgst"),
                "sgst":            _sum_items(items, "sgst"),
                "cess":            _sum_items(items, "cess"),
                "document_type":   doc_type,
                "is_amended":      False,
                "itc_availability": None,
                "return_period":   return_period,
                "place_of_supply": None,
                "reverse_charge":  False,
            }
            record["row_hash"] = _compute_2b_row_hash(record, salt)
            rows.append(record)
            row_num += 1

    df = pd.DataFrame(rows) if rows else _empty_gstr2b_dataframe()
    return df, errors


# ── Excel / CSV parser ─────────────────────────────────────────────────────────

# ── Large-file openpyxl streaming reader ──────────────────────────────────────

def _read_xlsx_streaming(
    raw_bytes: bytes,
    preferred_sheet: str = "B2B",
) -> pd.DataFrame:
    """
    Stream an xlsx workbook row-by-row using openpyxl read_only=True mode.

    Memory cost is O(1 row) instead of O(whole workbook), which is safe for
    GSTR-2B files larger than 5 MB (some companies have thousands of invoices).

    Tries 'preferred_sheet' by name first; falls back to the first sheet.
    Returns a string-typed DataFrame identical to pd.read_excel(..., dtype=str).

    Raises:
        ParseError: If the workbook cannot be opened or the sheet is empty.
    """
    buf = io.BytesIO(raw_bytes)
    try:
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    except Exception as exc:
        raise ParseError(f"openpyxl could not open xlsx workbook: {exc}") from exc

    try:
        ws = wb[preferred_sheet] if preferred_sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)

        # First row → header
        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            return pd.DataFrame()

        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]

        data = [
            [str(cell) if cell is not None else "" for cell in row]
            for row in rows_iter
        ]
    finally:
        wb.close()

    return pd.DataFrame(data, columns=headers)


def _parse_gstr2b_excel(
    raw_bytes: bytes,
    fmt: str,
    run_id: str,
    client_id: str,
    salt: str,
) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    """
    Parse a GSTR-2B Excel or CSV portal export.

    Normalizes column names using GSTR2B_EXCEL_ALIASES, then extracts
    the same canonical column set as the JSON parser.

    Handles the "B2B" sheet by name if present; falls back to the first sheet.

    Returns:
        (DataFrame with canonical columns, list of error dicts)
    """
    buf = io.BytesIO(raw_bytes)
    errors: list[dict[str, Any]] = []
    is_large = len(raw_bytes) > _LARGE_FILE_THRESHOLD_BYTES

    try:
        if fmt == "csv":
            df = pd.read_csv(buf, dtype=str)
        elif fmt == "xlsx":
            if is_large:
                # Stream rows without loading the entire workbook into RAM.
                logger.debug(
                    "parser.gstr2b.large_file_streaming",
                    size_mb=round(len(raw_bytes) / 1024 / 1024, 1),
                )
                df = _read_xlsx_streaming(raw_bytes, preferred_sheet="B2B")
            else:
                # Standard path: explicit engine="openpyxl" avoids auto-detection.
                try:
                    df = pd.read_excel(buf, sheet_name="B2B", engine="openpyxl", dtype=str)
                except Exception:
                    buf.seek(0)
                    df = pd.read_excel(buf, sheet_name=0, engine="openpyxl", dtype=str)
        else:  # xls — legacy format; xlrd does not support read_only streaming
            try:
                df = pd.read_excel(buf, sheet_name="B2B", engine="xlrd", dtype=str)
            except Exception:
                buf.seek(0)
                df = pd.read_excel(buf, sheet_name=0, engine="xlrd", dtype=str)
    except ParseError:
        raise
    except Exception as exc:
        raise ParseError(f"Failed to read GSTR-2B {fmt.upper()} file: {exc}") from exc

    # ── Column normalization ──────────────────────────────────────────────────
    # Build a lookup of lower-case raw column → canonical name
    col_map: dict[str, str] = {}
    raw_cols_lower = {c.strip().lower(): c for c in df.columns}

    for canonical, aliases in GSTR2B_EXCEL_ALIASES.items():
        for alias in aliases:
            if alias.lower() in raw_cols_lower:
                col_map[raw_cols_lower[alias.lower()]] = canonical
                break

    df = df.rename(columns=col_map)
    df = df.dropna(how="all")

    required = {"invoice_number", "gstin_supplier"}
    missing_required = required - set(df.columns)
    if missing_required:
        raise ParseError(
            f"GSTR-2B {fmt.upper()} missing required columns: {missing_required}. "
            f"Found columns after mapping: {sorted(df.columns.tolist())}"
        )

    rows: list[dict[str, Any]] = []

    for idx, row in df.iterrows():
        row_num = int(idx) + 2  # type: ignore[arg-type]  # +2 = 1-indexed + header

        gstin          = str(row.get("gstin_supplier", "")).strip().upper()
        invoice_number = str(row.get("invoice_number", "")).strip().upper()

        if not gstin or gstin in ("NAN", "NONE", ""):
            errors.append({"row": row_num, "reason": "Missing GSTIN — row skipped"})
            continue
        if not invoice_number or invoice_number in ("NAN", "NONE", ""):
            errors.append({"row": row_num, "reason": "Missing invoice number — row skipped"})
            continue

        # Parse document type from raw value
        raw_doc_type = str(row.get("document_type", "INV")).strip().upper()
        if raw_doc_type in ("C", "CR", "CREDIT NOTE", "CREDIT"):
            doc_type = DocumentType.CREDIT_NOTE
        elif raw_doc_type in ("D", "DR", "DEBIT NOTE", "DEBIT"):
            doc_type = DocumentType.DEBIT_NOTE
        else:
            doc_type = DocumentType.INVOICE

        record: dict[str, Any] = {
            "run_id":           run_id,
            "client_id":        client_id,
            "source_row":       row_num,
            "gstin_supplier":   gstin,
            "supplier_name":    str(row.get("supplier_name", "")).strip() or None,
            "invoice_number":   invoice_number,
            "invoice_date":     _parse_gst_date(row.get("invoice_date")),
            "taxable_value":    _coerce_numeric(row.get("taxable_value", 0)),
            "igst":             _coerce_numeric(row.get("igst", 0)),
            "cgst":             _coerce_numeric(row.get("cgst", 0)),
            "sgst":             _coerce_numeric(row.get("sgst", 0)),
            "cess":             _coerce_numeric(row.get("cess", 0)),
            "document_type":    doc_type,
            "is_amended":       False,
            "itc_availability": str(row.get("itc_availability", "")).strip() or None,
            "return_period":    str(row.get("return_period", "")).strip() or None,
            "place_of_supply":  str(row.get("place_of_supply", "")).strip() or None,
            "reverse_charge":   str(row.get("reverse_charge", "N")).strip().upper() == "Y",
        }
        record["row_hash"] = _compute_2b_row_hash(record, salt)
        rows.append(record)

    df_out = pd.DataFrame(rows) if rows else _empty_gstr2b_dataframe()
    return df_out, errors


# ── Empty-frame factory ────────────────────────────────────────────────────────

def _empty_gstr2b_dataframe() -> pd.DataFrame:
    """Return a zero-row DataFrame with the correct canonical columns."""
    return pd.DataFrame(columns=[
        "run_id", "client_id", "source_row",
        "gstin_supplier", "supplier_name",
        "invoice_number", "invoice_date",
        "taxable_value", "igst", "cgst", "sgst", "cess",
        "document_type", "is_amended", "itc_availability",
        "return_period", "place_of_supply", "reverse_charge",
        "row_hash",
    ])


# ── Public API ─────────────────────────────────────────────────────────────────

def parse_gstr2b(
    file_bytes: bytes,
    filename: str,
    salt: str,
    *,
    job_id: str,
    client_id: str,
    mime_type: str | None = None,
) -> ParseResult:
    """
    Parse a GSTR-2B file and return a ParseResult.

    Automatically detects whether the file is a JSON portal export or an
    Excel / CSV spreadsheet using the MIME type (preferred) and file
    extension (fallback), then routes to the appropriate parser.

    Args:
        file_bytes: Raw file content.
        filename:   Original filename (used for extension detection).
        salt:       Per-upload salt for deterministic row hashing.
        job_id:     Reconciliation job / run ID.
        client_id:  Tenant client UUID.
        mime_type:  Optional MIME type from the HTTP Content-Type header.

    Returns:
        ParseResult dataclass with records, totals, and errors.

    Raises:
        ParseError: On unrecognised format or fatal parse failure.
    """
    logger.info("parser.gstr2b.start", filename=filename, job_id=job_id, mime=mime_type)

    fmt = _detect_format(filename, mime_type)

    try:
        if fmt == "json":
            df, errors = _parse_gstr2b_json(file_bytes, job_id, client_id, salt)
        else:
            df, errors = _parse_gstr2b_excel(file_bytes, fmt, job_id, client_id, salt)
    except ParseError:
        raise
    except Exception as exc:
        raise ParseError(f"Failed to parse GSTR-2B '{filename}': {exc}") from exc

    total_rows  = len(df) + len(errors)
    parsed_rows = len(df)
    records = df.to_dict(orient="records") if not df.empty else []

    logger.info(
        "parser.gstr2b.complete",
        job_id=job_id,
        fmt=fmt,
        total=total_rows,
        parsed=parsed_rows,
        errors=len(errors),
    )

    return ParseResult(
        records=records,
        total_rows=total_rows,
        parsed_rows=parsed_rows,
        error_rows=len(errors),
        errors=errors,
    )
